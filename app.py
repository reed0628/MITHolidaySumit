import streamlit as st
import openpyxl
import random
import io
from datetime import datetime
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment

# --- 同事名單 ---
EMPLOYEE_LIST = [
    "陳育正 / Reed Chen", "蕭芮淇 / Charlotte Hsiao", "江亞璇 / Joyce Chiang",
    "陳幼慧 / Emily Chen", "高筑音 / Apple Kao", "林耕宇 / Benjamin", "林見松 / Jason Lin"
]

def get_random_time(sh, sm, eh, em):
    total_s, total_e = sh * 60 + sm, eh * 60 + em
    rnd = random.randint(total_s, total_e)
    return f"{rnd // 60:02d}:{rnd % 60:02d}"

# --- 升級版安全寫入：加入對齊與縮小字型功能 ---
def safe_write(ws, r, c, value, center=False, shrink=False, wrap=False):
    cell = ws.cell(row=r, column=c)
    target_cell = cell
    
    # 尋找合併儲存格的主格
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                target_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
                
    target_cell.value = value
    
    # 設定對齊格式 (保留原有格式，加上我們需要的新格式)
    current_align = target_cell.alignment
    target_cell.alignment = Alignment(
        horizontal='center' if center else current_align.horizontal,
        vertical='center' if center else current_align.vertical,
        shrink_to_fit=True if shrink else current_align.shrink_to_fit,
        wrap_text=True if wrap else current_align.wrap_text
    )

def process_excel(file):
    wb_read = openpyxl.load_workbook(file, data_only=True)
    wb_write = openpyxl.load_workbook(file)
    
    sheet_name = "海瀧簽到表" if "海瀧簽到表" in wb_write.sheetnames else wb_write.sheetnames[0]
    ws_read = wb_read[sheet_name]
    ws_write = wb_write[sheet_name]
    
    # 1. 寫入姓名 (B3) -> 開啟縮小以符合儲存格大小
    safe_write(ws_write, 3, 2, f"姓名：  {st.session_state.selected_name}", shrink=True)
    
    # 2. 自動尋找資料起始列
    start_row = 5
    for r in range(1, 10):
        if "序號" in str(ws_read.cell(row=r, column=1).value):
            start_row = r + 1
            break

    # 3. 處理出勤明細
    for row in range(start_row, start_row + 31):
        date_val = ws_read.cell(row=row, column=2).value
        desc_val = ws_read.cell(row=row, column=4).value
        
        is_empty_day = False
        if date_val is None or desc_val is None:
            is_empty_day = True
        elif isinstance(date_val, datetime) and date_val.year < 1905:
            is_empty_day = True
        elif str(date_val).strip() in ["", "0", "0.0", "None"] or str(desc_val).strip() in ["", "0", "0.0", "None"]:
            is_empty_day = True
            
        if is_empty_day:
            for col in range(1, 10):
                safe_write(ws_write, row, col, "")
            continue

        desc_str = str(desc_val).strip()
        try:
            if isinstance(date_val, datetime):
                date_str = date_val.strftime("%m/%d")
            elif "/" in str(date_val):
                date_str = str(date_val).strip()
            else:
                date_str = str(date_val)[5:10].replace("-", "/")
        except:
            date_str = ""

        # --- 假日畫 "--" 並置中 ---
        if "假日" in desc_str:
            for col in range(5, 10):
                safe_write(ws_write, row, col, "--", center=True)
            continue

        # --- 工作日填時間 ---
        if "工作" in desc_str:
            on_t = get_random_time(8, 50, 9, 5)
            off_t = get_random_time(18, 0, 18, 10)
            remark = ""

            if date_str in st.session_state.leaves:
                l = st.session_state.leaves[date_str]
                remark = f"{l['type']} {l['start']}-{l['end']}"
                if l['end'] == "12:00":
                    on_t = "13:30"
                elif l['start'] >= "13:30":
                    off_t = l['start']
                if l['start'] <= "09:00" and l['end'] >= "18:00":
                    on_t, off_t = "請假", "請假"

            safe_write(ws_write, row, 5, on_t)
            safe_write(ws_write, row, 6, "") # 清空簽到公式防 0
            safe_write(ws_write, row, 7, off_t)
            safe_write(ws_write, row, 8, "") # 清空簽退公式防 0
            # 備註欄開啟「縮小字型」與「自動換行」
            safe_write(ws_write, row, 9, remark, shrink=True, wrap=True)

    # --- 4. 【全域終極殺 0 行動】 ---
    # 掃描整張表，不管在哪個角落，只要公式算出來是 0，就把它清空！
    for r in range(1, ws_read.max_row + 1):
        for c in range(1, ws_read.max_column + 1):
            read_cell = ws_read.cell(row=r, column=c)
            # 如果讀到的值是 0 或 0.0
            if str(read_cell.value).strip() in ["0", "0.0"]:
                write_cell = ws_write.cell(row=r, column=c)
                # 直接清空該儲存格，消滅討厭的 0
                if not isinstance(write_cell, MergedCell):
                    write_cell.value = ""

    output = io.BytesIO()
    wb_write.save(output)
    return output.getvalue()

# --- 網頁介面 ---
st.set_page_config(page_title="海瀧出勤工具", layout="centered")
st.title("🚢 海瀧出勤紀錄自動填表")

st.session_state.selected_name = st.selectbox("1. 選擇姓名", EMPLOYEE_LIST)
uploaded_file = st.file_uploader("2. 上傳 Excel 範本", type=["xlsx"])

if uploaded_file:
    if 'leaves' not in st.session_state: st.session_state.leaves = {}
    st.subheader("3. 休假設定")
    c1, c2, c3, c4 = st.columns(4)
    with c1: d_in = st.text_input("日期(MM/DD)", placeholder="02/09")
    with c2: t_in = st.selectbox("假別", ["特休", "事假", "病假", "公假"])
    with c3: s_in = st.text_input("開始", "09:00")
    with c4: e_in = st.text_input("結束", "12:00")
    
    if st.button("➕ 新增休假"):
        if d_in:
            st.session_state.leaves[d_in] = {"type": t_in, "start": s_in, "end": e_in}
            st.rerun()

    if st.session_state.leaves:
        st.write("已設定休假：", st.session_state.leaves)
        if st.button("🗑️ 清空所有設定"):
            st.session_state.leaves = {}
            st.rerun()

    if st.button("🚀 生成並下載"):
        try:
            final_xlsx = process_excel(uploaded_file)
            download_name = st.session_state.selected_name.split(' / ')[0]
            st.download_button("💾 下載成果 Excel", final_xlsx, f"{download_name}_出勤表.xlsx")
        except Exception as e:
            st.error(f"錯誤：{e}")
